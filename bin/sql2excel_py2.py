#!/usr/bin/python
# -*- coding: UTF-8 -*-
#coding=utf-8
"""
    输入一个sql文件，转换为excel文件
    sql文件放在../input下 sql文件格式要是utf-8的
    生成的excel放在../output下
    在../bin下执行 python sql2excel.py xxx.sql(直接文件名 不需要路径)
    V1.0.2_20181220
"""
import sys
import re
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
#########################################
sys_encoding = sys.getfilesystemencoding()
reload(sys)
sys.setdefaultencoding("utf-8")
#########################################
def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill
######################################################################
font =Font(name='Vrinda')
font2=Font(name='Vrina',bold=True)#粗体

fill_green = PatternFill("solid", fgColor="C6E0B4")
fill_blue  = PatternFill("solid", fgColor="9BC2E6")
fill_yellow= PatternFill("solid", fgColor="FFE699")
fill_pink  = PatternFill("solid", fgColor="FF99CC")

border = Border(top   = Side(border_style="thin",color='000000'),
                left  = Side(border_style="thin",color='000000'), 
                right = Side(border_style="thin",color='000000'), 
                bottom= Side(border_style="thin",color='000000'))

link = "test.xlsx#Sheet!A1"

######################################################################
workbook=Workbook()
worksheet=workbook.active
######################################################################

input_file=open('../input/'+sys.argv[1],'r')
flag=0#1表示找到了表明 2表示进入（ 
table_all      ={}
table_sturct   ={}
current_comment=''

while 1:
    line=input_file.readline()
    if not line:
        break
    line=line.strip()
    #print(line.decode('utf-8').encode(sys_encoding))

    if re.match(r'^create table',line):
        if re.search(r'^create table\s.*\.\w+',line,re.M|re.I):
            table_name=re.search(r'^create table\s.*\.(\w+)',line,re.M|re.I).group(1)
        else:
            table_name=re.search(r'^create table\s*(\w+)',line,re.M|re.I).group(1)
        if not table_all.has_key(table_name):
            #print(table_name)
            table_all[table_name]=OrderedDict()#保证输入的时候进去的是什么顺序，输出的时候也是什么顺序

        flag=1

    if re.match(r'^\($',line) and flag==1:
        flag=2
        table_sturct=[]
        continue

    if re.match(r'^\);*$',line) and flag==2:
        flag=0
        continue

    if flag==2:
        searchObj = re.search( r'^([^\s]+)\s+([^\s]+),*', line, re.M|re.I)
        table_all[table_name][searchObj.group(1)]=[]
        table_all[table_name][searchObj.group(1)].append(searchObj.group(2))
    
    if re.match(r'^comment on column [\w\d_]+\.[\w\d_]+\.[\w\d_]+',line):#获取comment上的表名字 db_name.table_name.field_name
        searchObj = re.search( r'comment on column [\w\d_]+\.[\w\d_]+\.([\w\d_]+)', line, re.M|re.I)
        current_field=searchObj.group(1)
        flag=3
        #print line

    if flag==3 and re.match(r'\s*is.+\';$',line):#匹配到commit的说明且commt是单行的
        flag=0
        searchObj=re.search(r'\s*is\s+\'(.*)\';', line, re.M|re.I) 
        table_all[table_name][current_field].append(searchObj.group(1))
    elif flag==3 and re.match(r'\s*is.+[^;]$',line):#如果comments分行了，此处是第一行
        searchObj=re.search(r'\s*is\s+\'(.*)\'*', line, re.M|re.I) 
        current_comment=searchObj.group(1)
    elif flag==3 and re.search(r';$',line):#如果comments分行了，此处是最后一行
        searchObj=re.search(r'(.*)\'*;', line, re.M|re.I) 
        current_comment+=' '
        current_comment+=searchObj.group(1)
        table_all[table_name][current_field].append(current_comment)
        flag=0
    elif flag==3:#如果comments分行了，此处是中间行
        current_comment+=' '
        current_comment+=line

input_file.close()

table_num      =2
for tab_name in table_all:  
    ws_tep=workbook.create_sheet(tab_name)
    worksheet['A'+str(table_num)]=tab_name
    worksheet['A'+str(table_num)].hyperlink = ("#"+tab_name+"!A1")
    ws_tep['A1']='Return'
    ws_tep['A1'].font=font
    ws_tep['A1'].fill=fill_green
    ws_tep['A1'].hyperlink=("#Sheet!A"+str(table_num))
    table_num+=1
    ws_tep['A3']=tab_name
    merge_cell=ws_tep.merge_cells('A3:B3')
    style_range(ws_tep,'A3:B3',border,fill_yellow,font2,None)
    ws_tep['A4']='No.'
    ws_tep['A4'].font=font2
    ws_tep['A4'].fill=fill_blue
    ws_tep['A4'].border=border

    ws_tep['B4']='File Name'
    ws_tep['B4'].font=font2
    ws_tep['B4'].fill=fill_blue
    ws_tep['B4'].border=border

    ws_tep['C4']='File Type'
    ws_tep['C4'].font=font2
    ws_tep['C4'].fill=fill_blue
    ws_tep['C4'].border=border

    ws_tep['D4']='Description'
    ws_tep['D4'].font=font2
    ws_tep['D4'].fill=fill_blue
    ws_tep['D4'].border=border

    ws_tep['E4']='Remark'
    ws_tep['E4'].font=font2
    ws_tep['E4'].fill=fill_blue
    ws_tep['E4'].border=border

    filed_num=5
    for fed_name in table_all[tab_name]:
        ws_tep['A'+str(filed_num)]=filed_num-4
        ws_tep['A'+str(filed_num)].font=font
        ws_tep['A'+str(filed_num)].fill=fill_green
        ws_tep['A'+str(filed_num)].border=border

        ws_tep['B'+str(filed_num)]=fed_name
        ws_tep['B'+str(filed_num)].font=font
        ws_tep['B'+str(filed_num)].fill=fill_yellow
        ws_tep['b'+str(filed_num)].border=border

        ws_tep['C'+str(filed_num)]=table_all[tab_name][fed_name][0]
        ws_tep['C'+str(filed_num)].font=font
        ws_tep['C'+str(filed_num)].fill=fill_pink
        ws_tep['C'+str(filed_num)].border=border

        if len(table_all[tab_name][fed_name])==1:            
            ws_tep['D'+str(filed_num)]=''
        else:
            print(table_all[tab_name][fed_name][1])  #xxx..decode('utf-8').encode(sys_encoding))
            table_all[tab_name][fed_name][1]=table_all[tab_name][fed_name][1]
            ws_tep['D'+str(filed_num)]=table_all[tab_name][fed_name][1]
        ws_tep['D'+str(filed_num)].font=font
        ws_tep['D'+str(filed_num)].fill=fill_pink
        ws_tep['D'+str(filed_num)].border=border

        ws_tep['E'+str(filed_num)]=""
        ws_tep['E'+str(filed_num)].font=font
        ws_tep['E'+str(filed_num)].fill=fill_pink
        ws_tep['E'+str(filed_num)].border=border

        filed_num+=1

workbook.save('../output/'+sys.argv[1]+'.xlsx') 